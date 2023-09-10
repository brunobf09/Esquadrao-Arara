# Import
import streamlit as st
import func as f
import pandas as pd
import datetime
from openpyxl.styles import Border, Side, Alignment
import time

st.set_page_config(page_title="Esquadrão Arara")

# Sidebar
with st.sidebar:
    #Style
    st.markdown("<h1 style='text-align: center; "
                "margin-top:-80px; color: "
                "black;'"
                ">Tudo Sob Nossas Asas!</h1>", unsafe_allow_html=True)
    st.markdown( """
        <style>
            [data-testid=stSidebar] [data-testid=stImage]{
                text-align: center;
                display: block;
                margin-top:-55px;
                margin-left: auto;
                margin-right: auto;
                width: 50%;
            }
        </style>
        """, unsafe_allow_html=True)
    #content
    st.image("Pic/bolacha.png")
    pages = st.selectbox('Selecione a página desejada', ['','Escala de Voo','Planejamento de Missão','Ordem de Missão'])

#landpage
if pages == '':
    #Style
    st.markdown(f.landpage(), unsafe_allow_html=True)
    st.markdown("<h1 style= margin-top:-100px;>"
                "Bem-vindo ao Esquadrão Arara</h1>",unsafe_allow_html=True)
    st.markdown('<p style= margin-top:-45px;> Por: Bruno Brasil</p>',unsafe_allow_html=True)

#Page - Escala
if pages == "Escala de Voo":
    # Page desing
    st.title('**Escala de Voo**')

    #options box
    options = st.selectbox('Selecione a função desejada:',["","Disponibilidade","Pau de Sebo", "Quadrinhos","Escala"])

    calendario = {1:"JANEIRO", 2:"FEVEREIRO", 3:"MARÇO",4:"ABRIL",5:"MAIO",
                      6:"JUNHO",7:"JULHO",8:"AGOSTO",9:"SETEMBRO",10:"OUTUBRO",
                      11:"NOVEMBRO",12:"DEZEMBRO"}

    if options == "Disponibilidade":
        inicio = st.date_input('Início da Disponibilidade',format='DD/MM/YYYY')
        fim = st.date_input('Término da Disponibilidade',format="DD/MM/YYYY")
        mes = calendario.get(inicio.month)

        if fim >= inicio:
            try:
                st.table(f.indisp(inicio.day,fim.day,mes))
            except:
                pass

    if options == "Pau de Sebo":
        sebo = f.sebo()
        sebo = sebo.style.applymap(f.desadapt, subset=['Último Voo'])
        st.table(sebo)

    if options == "Quadrinhos":
        labels = f.label_quad()
        quadrinho = st.selectbox('Selecione o quadrinho desejado:', labels)

        #Filtro por operacionalidade
        if quadrinho!="":
            funcao = st.multiselect('Filtrar por função a bordo:',
                                ['PILOTO', 'MECÂNICO', 'LOADMASTER'],
                                ['PILOTO', 'MECÂNICO', 'LOADMASTER'])

            op = st.multiselect('Filtrar por operacionalidade:',
                                ['IN', 'OP', 'PB', 'AL'],
                                ['IN', 'OP', 'PB', 'AL'])

        #checkbox para verificar disponibilidade
        disp = st.checkbox('DISPONIBILIDADE')

        try:
            #Quadrinho + disponibilidade
            if disp:
                inicio = st.date_input('Início da Disponibilidade',format="DD/MM/YYYY")
                fim = st.date_input('Término da Disponibilidade',format="DD/MM/YYYY")
                mes = calendario.get(inicio.month)
                ind = f.indisp_quad(inicio.day, fim.day, mes)
                df = f.quad(quadrinho,funcao,op)
                st.table(df[df.index.isin(ind, level=1)])
            # Somente Quadrinho
            else:
                st.table(f.quad(quadrinho,funcao,op))
        except:
            pass

    if options == "Escala":
        #Quadrinhos
        labels = f.label_quad()
        quadrinho = st.selectbox('Selecione o quadrinho desejado:', labels)

        if quadrinho!="":
            #Datas
            inicio = st.date_input('Início da Disponibilidade',format="DD/MM/YYYY")
            fim = st.date_input('Término da Disponibilidade',format="DD/MM/YYYY")
            mes = calendario.get(inicio.month)
            mes_plan = inicio.month
            #Disponibilidades
            ind = f.indisp_quad(inicio.day, fim.day, mes)

            #Planejamento de horas de voo
            df = f.plan(mes_plan)
            quad = f.quad(quadrinho, 'PILOTO', ['IN', 'OP', 'PB', 'AL'])
            quad = quad.reset_index().set_index(1)
            df = quad.join(df).sort_values(by=['Meta','Quadrinhos'], ascending=[False,True]).drop(0,axis=1)
            sebo = f.sebo()
            df = df.join(sebo)
            prioridade = df[df.index.isin(ind, level=0)]
            prioridade = prioridade.reset_index().set_index(9) #rescrever depois
            prioridade.columns = ["Pilotos", "Quadrinhos", 'Meta', 'Horas Voadas','Último Voo']
            prioridade = prioridade[["Pilotos", "Meta", 'Quadrinhos', 'Horas Voadas','Último Voo']]
            prioridade.Meta = prioridade.Meta.astype('int')
            prioridade['Último Voo'] = prioridade['Último Voo'].astype('int')
            prioridade.drop('Horas Voadas', axis=1, inplace=True)


            try:
                st.subheader('Instrutores')
                IN = prioridade.loc['IN'].set_index("Pilotos")
                IN = IN.style.applymap(f.desadapt, subset=['Último Voo'])
                st.table(IN)
            except KeyError:
                st.write("Nenhum piloto disponível.")

            try:
                st.subheader('Operacionais')
                OP = prioridade.loc['OP'].set_index("Pilotos")
                OP = OP.style.applymap(f.desadapt, subset=['Último Voo'])
                st.table(OP)
            except KeyError as err:
                st.write("Nenhum piloto disponível.")

            try:
                st.subheader('Básicos')
                PB = prioridade.loc['PB'].set_index("Pilotos")
                PB = PB.style.applymap(f.desadapt, subset=['Último Voo'])
                st.table(PB)
            except KeyError as err:
                st.write("Nenhum piloto disponível.")

            try:
                st.subheader('Alunos')
                AL = prioridade.loc['AL'].set_index("Pilotos")
                AL = AL.style.applymap(f.desadapt, subset=['Último Voo'])
                st.table(AL)
            except KeyError as err:
                st.write("Nenhum piloto disponível.")

#Page - Plan
if pages == "Planejamento de Missão":
    #Título da página
    st.title('**Planejamento de Missão**')

    #sidebar
    with st.sidebar:
        n_days = st.number_input('Número de Dias', 1)
        tripul = st.number_input('Tripulantes',5)
        aeronave = st.selectbox("Selecione a Aeronave", ['FAB2802','FAB2805','FAB2809'])
        pb = {'FAB2800':12924,'FAB2802':12584,'FAB2803':12587,'FAB2805':12227,'FAB2809':12446}
        PBO = pb.get(aeronave) + 300 + (tripul*100)


    #inputs
    labels = f.data_icao_label()
    i = 0
    total_plan=[]
    while i < n_days:
        #inputs boxes
        data = st.date_input('Início da Missão', key=f'data_{i}', format="DD/MM/YYYY")
        hora = st.time_input('Horário da Decolagem', datetime.time(12, 00), key=f'hora_{i}', step=300)
        rota = st.multiselect("Selecione a Rota",labels, key=f'rota_{i}')
        try:
            alternativa = st.multiselect("Selecione as Alternativas", labels, max_selections=(len(rota) - 1),
                                     key=f'alt_{i}')
        except:
            break
            
        noabast = st.multiselect("Não Abastece",labels, key=f'abast_{i}')

        #plan fuction
        plan = f.braplan(data, hora, rota, alternativa)
        plan = f.disp(plan,noabast, PBO)
        total_plan.append(plan)

        #table check
        st.table(plan)
        checked = st.checkbox('Checked', key=f'checked_{i}')

        if checked:
            i+=1
            checked=False
        else:
            checked=False
            break

    #End Plan
    def edit_save():
            if st.session_state.save:
                st.session_state["final_data"] = edit_plan
                st.session_state.edit = False
    if i==n_days:
        st.title("Planejamento Completo")
        edit = st.checkbox('Editar', key='edit')
        end_plan = pd.concat(total_plan)

        if edit:
            if "final_data" not in st.session_state:
                edit_plan = end_plan.astype(str)
                edit_plan = st.data_editor(edit_plan, hide_index=True, on_change=edit_save)
                save = st.button("Salvar alteração", key='save', on_click=edit_save)
            else:
                edit_plan = st.session_state["final_data"].astype(str)
                edit_plan = st.data_editor(edit_plan, hide_index=True, on_change=edit_save)
                save = st.button("Salvar alteração", key='save', on_click=edit_save)

        elif edit==False:
            if "final_data" in st.session_state:
                end_plan = st.session_state["final_data"]
                st.table(end_plan)
            else:
                st.table(end_plan)

            # Download button
            csv = f.convert_df(end_plan)
            st.download_button(
                label="Download Excel",
                data=csv,
                file_name='Araraplan.csv',
                mime='text/csv',
            )

# Page - Ordem de missão
if pages == "Ordem de Missão":
    # Título da página
    st.title('**Ordem de Missão**')
    # Config
    thin = Side(border_style="thin", color="000000")

    # Inputs OM
    arquivo_upload = st.file_uploader("Faça upload do planejamento", type=["csv"])
    efetivo = f.efetivo()

    col1, col2, col3 = st.columns(3)
    with col1:
        om_number = st.text_input("Informe o número da OM")
    with col2:
        ofrag = st.text_input("Informe o número da OFRAG")
    with col3:
        missao = st.text_input("Informe o tipo da missão")

    col1, col2, col3 = st.columns(3)
    with col1:
        esf_aer = st.text_input("Informe o Esforço Aéreo (ind S/N)")
    with col2:
        caixa_nav = st.text_input("Informe a caixa de navegação")
    with col3:
        spot = st.text_input("Informe o número do SPOT")

    col1, col2, col3 = st.columns(3)
    with col1:
        pil = st.multiselect('Selecione os Pilotos:', efetivo.index, key="pil")
    with col2:
        mec = st.multiselect('Selecione os Mecânicos:', efetivo.index, key="mec")
    with col3:
        lm = st.multiselect('Selecione os Loadmaster:', efetivo.index, key="lm")

    aeronave = st.selectbox("Selecione a Aeronave:", ['', 'FAB2802', 'FAB2805', 'FAB2809'])
    detalhamento = st.text_area("Detalhamento da missão:")

    if arquivo_upload is not None:
        # carregar arquivos
        efetivo = f.efetivo()
        df = pd.read_csv(arquivo_upload, sep=',')

        # Inicializador
        gerar_om = st.button("Gerar OM")
        if gerar_om:
            wb = f.workload()
            om = wb['Ordem de Missão']
            om['C5'].value = om_number + '/1GAV9/2023'
            om['C7'].value = efetivo.loc[pil[0]].guerra
            om['I5'].value = df.DATA[0]
            om['A16'].value = detalhamento
            om['B26'].value = missao
            om['E26'].value = ofrag
            om['J26'].value = caixa_nav
            om['C28'].value = esf_aer
            om['J28'].value = spot
            om['J7'].value = f.serie_to_timedelta(df.TEV, sum=True)
            om['C9'].value = aeronave
            om['X1'].value = "/ ".join(pil + mec + lm)
            om['X2'].value = missao

            trip = pil + mec + lm
            zimbra = f.zimbra(efetivo, trip)
            om['O2'] = zimbra
            om['O3'] = f'OM{om_number} {missao}'

            pil = f.trigname(efetivo, pil)
            mec = f.trigname(efetivo, mec)
            lm = f.trigname(efetivo, lm)

            for n, p in enumerate(pil):
                om.insert_rows(11 + n)
                om.cell(11 + n, 1).value = p

            for n, m in enumerate(mec):
                om.insert_rows(12 + n + len(pil))
                om.cell(12 + n + len(pil), 1).value = m

            for n, l in enumerate(lm):
                om.insert_rows(13 + n + len(pil) + len(mec))
                om.cell(13 + n + len(pil) + len(mec), 1).value = l

            sum_len = len(pil) + len(mec) + len(lm)

            # plan
            len_plan = len(df)
            for n in range(len_plan):
                om.insert_rows(15 + n + sum_len)
                for i in range(10):
                    om.cell(15 + n + sum_len, i + 1).value = df.iloc[n, i]
                    om.cell(15 + n + sum_len, i + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    om.cell(15 + n + sum_len, i + 1).alignment = Alignment(horizontal='center')

            line1 = 16 + sum_len + len_plan
            line2 = 24 + sum_len + len_plan
            merge = f'A{line1}:J{line2}'
            om.merge_cells(merge)
            # om.row_dimensions[line].height = 160

            # Pedido de lanche
            lanche = wb['Lanche de Bordo']
            tripulacao = pil + mec + lm

            # Inputs
            lanche['A16'] = f'DESTINO: {df.ARR[0]}'
            lanche['AZ14'] = df['HORA(Z)'][0]
            lanche['A20'] = f'Qtd de Lanches: {len(tripulacao)}'

            for n, t in enumerate(tripulacao):
                lanche.insert_rows(25 + n)
                lanche.cell(25 + n, 1).value = t
                lanche.cell(25 + n, 6).value = 'BAMN'
                lanche.cell(25 + n, 6).alignment = Alignment(horizontal='center')
                lanche.cell(25 + n, 7).value = 1
                lanche.cell(25 + n, 7).alignment = Alignment(horizontal='center')
                lanche.cell(25 + n, 8).value = 0
                lanche.cell(25 + n, 8).alignment = Alignment(horizontal='center')

            with st.spinner('Confeccionando OM...'):
                time.sleep(3)

            st.success("Ordem de Missão pronta para download!")
            # Criar um botão no Streamlit para baixar o arquivo
            bytes_data = f.excel_to_bytes(wb)
            st.download_button('Baixar OM', data=bytes_data, file_name=f'OM{om_number} {missao}.xlsx',
                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')























